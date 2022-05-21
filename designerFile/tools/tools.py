import csv
import codecs
import datetime

import openpyxl
from PyQt5.QtWidgets import QInputDialog
from openpyxl import load_workbook
from pandas.io.excel import ExcelWriter
import pandas as pd
import xlsxwriter
import xlrd as xlrd
import os
import re
import win32file


def getblockNmaeListfromcsv(Path):
    with open(Path, "r") as csvfile:
        reader = csv.reader(csvfile)
        rows = [row for row in reader]
        blockNmaeList = []
        for j in range(1, len(rows[0])):
            if (rows[0][j] and rows[0][j] != ' '):
                blockNmaeList.append(rows[0][j])
        return blockNmaeList


def getblockNmaeListfromxlsx(xlsxpath):
    blockNmaeList = []
    # 打开文件
    workbook = xlrd.open_workbook(xlsxpath)
    # 根据sheet索引或者名称获取sheet内容
    sheet = workbook.sheet_by_index(0)  # sheet索引从0开始
    cols = sheet.ncols  # 获取有多少列
    for j in range(cols):
        value = sheet.cell_value(0, j)
        if value:
            blockNmaeList.append(value)
    return blockNmaeList


class Data():
    le = 0
    ri = 0
    GVWindexList = []


def getxlsxData(path, blockName):
    list = path.rsplit("/", 1)
    print('---------------------------', list)
    filename = list[1]
    filepath = list[0]
    # if not re.search('.xlsx', filename):
    #     xlsxname = filename.split('.')[0] + '.xlsx'
    #     xlsxPath = os.path.join(filepath, xlsxname)
    #     workbook = xlsxwriter.Workbook(xlsxPath)
    #     worksheet = workbook.add_worksheet('first_sheet')
    #     workbook.close()
    #     with ExcelWriter(xlsxPath) as ew:
    #         pd.read_csv(path).to_excel(ew, index=False)
    # else:
    xlsxname = filename
    demoxlsxpath = readData(filepath, xlsxname, blockName)
    ans = [Data.le, Data.ri, demoxlsxpath, Data.GVWindexList]
    return ans


def readData(filepath, xlsxname, blockName):
    xlsxpath = os.path.join(filepath, xlsxname)
    # 打开文件
    workbook = xlrd.open_workbook(xlsxpath)
    # 根据sheet索引或者名称获取sheet内容
    sheet = workbook.sheet_by_index(0)  # sheet索引从0开始
    rows = sheet.nrows  # 获取有多少行
    cols = sheet.ncols  # 获取有多少列
    blocknAME = blockName
    # sheet.cell_value(第几行,第几列)
    demoxlsxpath = os.path.join(filepath, 'cache.xlsx')
    workbook = xlsxwriter.Workbook(demoxlsxpath)  # 创建一个excel文件
    # 在文件中创建一个名为TEST的sheet,不加名字默认为sheet1
    worksheet = workbook.add_worksheet(u'sheet1')
    for i in range(0, rows):
        for j in range(0, 4):
            value = sheet.cell_value(i, j)
            if (re.search('Unnamed', str(value))):
                value = ''
            worksheet.write(i, j, value)

    for j in range(4, cols):
        value = sheet.cell_value(1, j)
        if (value == 'GVW'):
            Data.GVWindexList.append(j)
    left = 0
    right = 0
    flag = 0
    for l in range(4, cols):
        if (flag == 1):
            break
        newvalue = sheet.cell_value(0, l)
        if (newvalue == blocknAME):
            left = l
            flag = 1
            for r in range(l + 1, cols):
                value = sheet.cell_value(0, r)
                searchObj = re.search('Unnamed', value, re.M | re.I)
                if (not searchObj and value):
                    right = r
                    break
                right = cols
    Data.le = left
    Data.ri = right
    for i in range(0, rows):
        x = 4
        for j in range(left, right):
            value = sheet.cell_value(i, j)
            if (re.search('Unnamed', str(value))):
                value = ''
            worksheet.write(i, x, value)
            x += 1

    workbook.close()
    return demoxlsxpath


def csv_to_xlsx_pd(cacheFilepath, filepath, filename):
    csvPath = os.path.join(filepath, filename)
    xlsxname = filename.split('.')[0] + '.xlsx'
    xlsxPath = os.path.join(cacheFilepath, 'cache', xlsxname)
    workbook = xlsxwriter.Workbook(xlsxPath)
    worksheet = workbook.add_worksheet('first_sheet')
    workbook.close()
    with ExcelWriter(xlsxPath) as ew:
        pd.read_csv(csvPath).to_excel(ew, index=False)

    return xlsxname


def csv_to_xlsx(filepath, csvPath):
    datetime_object = datetime.datetime.now()
    time = str(datetime_object).split(' ')[0] + '-' + str(datetime_object).split('.')[1]
    xlsxname = 'Load_ IterationN' + time + '.xlsx'
    xlsxPath = os.path.join(filepath, xlsxname)
    workbook = xlsxwriter.Workbook(xlsxPath)  # 创建一个excel文件
    worksheet = workbook.add_worksheet(u'Sheet1')  # 在文件中创建一个名为TEST的sheet,不加名字默认为sheet1
    with open(csvPath, 'r', encoding='utf-8') as f:
        read = csv.reader(f)
        l = 0
        for line in read:
            r = 0
            for i in line:
                if (re.search('noname', i)):
                    i = ''
                else:
                    worksheet.write(l, r, i)
                r = r + 1
            l = l + 1
    workbook.close()
    return xlsxname


def addData2xlsx(xlsxPathcache, xlsxPath):
    workbook = xlrd.open_workbook(xlsxPathcache)
    sheet = workbook.sheet_by_index(0)  # sheet索引从0开始
    rows = sheet.nrows  # 获取有多少行
    cols = sheet.ncols  # 获取有多少列
    datetime_object = datetime.datetime.now()
    time = str(datetime_object).split(' ')[0] + '-' + str(datetime_object).split('.')[1]
    SheetName = 'Load_iterationN' + time
    print(SheetName)
    wb = openpyxl.load_workbook(xlsxPath)
    wb.create_sheet(SheetName)
    dest_sheet = wb.get_sheet_by_name(SheetName)
    for i in range(0, rows):
        for j in range(0, cols):
            value = str(sheet.cell_value(i, j))
            if (re.search('Unnamed', value)):
                value = ''
            if (j >= 26):
                t = int(j) - 26
                x = chr(t + 65)
                x = 'A' + x
                item = x + str(i + 1)
                dest_sheet[item] = value
            else:
                x = chr(int(j) + 65)
                item = x + str(i + 1)
                dest_sheet[item] = value
    wb.save(xlsxPath)


def washXlsx(xlsxPath):
    workbook = xlrd.open_workbook(xlsxPath)
    sheet = workbook.sheet_by_index(0)  # sheet索引从0开始
    rows = sheet.nrows  # 获取有多少行
    cols = sheet.ncols  # 获取有多少列
    wb = openpyxl.load_workbook(xlsxPath)
    dest_sheet = wb.get_sheet_by_name(wb.sheetnames[0])
    for i in range(0, rows):
        for j in range(0, cols):
            value = str(sheet.cell_value(i, j))
            if (re.search('Unnamed', value)):
                value = ''
            if (j >= 26):
                t = int(j) - 26
                x = chr(t + 65)
                x = 'A' + x
                item = x + str(i + 1)
                dest_sheet[item] = value
            else:
                x = chr(int(j) + 65)
                item = x + str(i + 1)
                dest_sheet[item] = value
    wb.save(xlsxPath)


def xlsx_to_csv(xlsxpath):
    list = xlsxpath.rsplit("/", 1)
    filename = list[1]
    filepath = list[0]
    csvname = filename.split('.')[0] + '.csv'
    csvPath = os.path.join(filepath, csvname)
    workbook = xlrd.open_workbook(xlsxpath)
    table = workbook.sheet_by_index(0)
    with codecs.open(csvPath, 'w', encoding='utf-8') as f:
        write = csv.writer(f)
        for row_num in range(table.nrows):
            row_value = table.row_values(row_num)
            for i in range(len(row_value)):
                if row_value[i]:
                    if (re.search('Unnamed', str(row_value[i]))):
                        row_value[i] = ''
            write.writerow(row_value)


def is_used(file_name):
    try:
        v_handle = win32file.CreateFile(file_name, win32file.GENERIC_READ, 0, None, win32file.OPEN_EXISTING,
                                        win32file.FILE_ATTRIBUTE_NORMAL, None)
        result = bool(int(v_handle) == win32file.INVALID_HANDLE_VALUE)
        win32file.CloseHandle(v_handle)
    except Exception:
        return True
    return result


def ScientificEnumeration2Number(a):
    li = a.split('e')
    x = 0
    if (len(li) == 2):
        e = float(li[0])
        f = int(li[1])
        x = e * pow(10, f)
    else:
        x = float(li[0])
    return x


def ScientificEnumerationFormatting(a):
    li = str(a).split('e')
    x = 0
    if (len(li) == 2):
        e = float(li[0])
        f = int(li[1])
        x = e * pow(10, f)
    else:
        x = float(li[0])
    return f"{x:.2e}"


def copyXlsx(xlsxpath, cacheFilepath, flag):
    workbook = xlrd.open_workbook(xlsxpath)
    sheet = workbook.sheet_by_index(0)  # sheet索引从0开始
    rows = sheet.nrows  # 获取有多少行
    cols = sheet.ncols  # 获取有多少列
    if flag == 0:
        demoxlsxpath = os.path.join(cacheFilepath, 'cache', 'loadPathcache.xlsx')
    else:
        datetime_object = datetime.datetime.now()
        time = str(datetime_object).split(' ')[0] + str(datetime_object).split('.')[1]
        xlsxname = 'Load_ IterationN' + time + '.xlsx'
        demoxlsxpath = os.path.join(cacheFilepath, 'cache', xlsxname)
    workbook2 = xlsxwriter.Workbook(demoxlsxpath)  # 创建一个excel文件
    # 在文件中创建一个名为TEST的sheet,不加名字默认为sheet1
    worksheet = workbook2.add_worksheet(u'Sheet1')
    for i in range(0, rows):
        for j in range(0, cols):
            value = sheet.cell_value(i, j)
            worksheet.write(i, j, value)
    workbook2.close()
    return demoxlsxpath.replace('\\', '/')
