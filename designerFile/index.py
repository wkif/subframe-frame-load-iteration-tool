# -*- coding: utf-8 -*-
# author:kif<kif101001000@163.com>
# time: 2022,04,29
# Form implementation generated from reading ui file 'connect_me.ui'
#
# Created by: PyQt5 UI code generator 5.11.3
#
# WARNING! All changes made in this file will be lost!
# 导入程序运行必须模块
import datetime
import os
import re
import shutil
import sys
import time
from decimal import Decimal
import QCandyUi
from threading import Timer
import openpyxl
import xlsxwriter
from QCandyUi import CandyWindow
from openpyxl import load_workbook
import xlrd
from PyQt5 import QtWidgets, QtCore
from PyQt5.QtCore import QDir, Qt, QVersionNumber, QT_VERSION_STR
from PyQt5.QtWidgets import *
from PyQt5.QtGui import QIcon, QBrush, QColor, QFont
from openpyxl.styles import PatternFill
from qt_material import apply_stylesheet

import pandas as pd
import numpy as np

from designerFile.tools.help import getHelp
from designerFile.tools.log import createLog
from designerFile.tools.sendLog import sendLog
from designerFile.tools.tools import getxlsxData, csv_to_xlsx_pd, csv_to_xlsx, xlsx_to_csv, getblockNmaeListfromcsv, \
    is_used, \
    ScientificEnumeration2Number, ScientificEnumerationFormatting, addData2xlsx, washXlsx, getblockNmaeListfromxlsx, \
    copyXlsx
from designerFile.mainView import Ui_MainWindow
from designerFile.view2 import Ui_Dialog


class ChildWin_block(QtWidgets.QDialog, Ui_Dialog):
    def __init__(self):
        super(ChildWin_block, self).__init__()
        self.setupUi(self)


class helpWindow(QMainWindow):
    def __init__(self):
        # super().__init__()
        super(helpWindow, self).__init__()
        self.setWindowTitle('帮助')
        layout = QVBoxLayout()

        file_text = getHelp()
        label = QLabel(self)
        label.setFrameStyle(QFrame.Panel | QFrame.Sunken)
        label.setText(file_text)
        # label.setAlignment(Qt.AlignBottom | Qt.AlignRight)
        layout.addWidget(label)
        collec_btn = QPushButton('提交日志', self)
        collec_btn.clicked.connect(sendLog)
        layout.addWidget(collec_btn)
        widget = QWidget()
        widget.setLayout(layout)
        self.setCentralWidget(widget)

        # self.resize(280, 230)

        #


class globalData():
    HotSpots_N_List = []
    bolckNameList = []
    LoadDatapath = ''
    VRLDApath = ''
    LoadDatapath_xlsx = ''
    userFilepath = ''
    BlockCount = 0
    firstFlag = 0
    left = 0
    right = 0
    editChangeFlag = 0
    cacheFilepath = ''
    GVWflag = 0,
    t = None,
    cycleFlag = 0
    GVWindexList = []
    blockIndex = 0
    dialog = ''
    scrollBar_A = ''
    scrollBar_B = ''
    scrollBar_C = ''
    scrollBar_A2 = ''
    scrollBar_B2 = ''
    scrollBar_C2 = ''
    save_loadDataPath = ''
    save_resultDataPath = ''
    loadDataisXLSX = False
    VRLDAisXLSX = False
    temploadDataPath = ''
    writeXlsxPath = ''
    Result_IterationName = ''
    updateFlag = True
    templateFlag = False


class MyMainForm(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super(MyMainForm, self).__init__(parent)
        self.setupUi(self)
        # 按钮事件绑定
        self.setWindowIcon(QIcon('F:/File_my/Project/Plug-in_1/file/logo.png'))
        self.pushButton.clicked.connect(self.ImportVRLDA)
        self.pushButton_2.clicked.connect(self.ImportLoadData)
        self.pushButton_3.clicked.connect(self.ImportCYCLE)
        self.pushButton_4.clicked.connect(self.noticeSave)
        self.pushButton_5.clicked.connect(self.writeXlsx)
        self.pushButton_9.clicked.connect(self.clearAll)
        # self.pushButton_10.clicked.connect(self.helpMessage)
        self.pushButton_11.clicked.connect(self.importTemp)
        self.tableWidget_2.horizontalHeader().sectionClicked.connect(self.HorSectionClicked)  # 表头单击信号
        # tableWidget
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # 左表格铺满整个QTableWidget控件
        self.tableWidget.verticalHeader().setVisible(False)
        # 隐藏列标题
        self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        # 隐藏竖直滚动条
        self.tableWidget.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)
        # 中表格不可编辑
        # font = self.tableWidget.horizontalHeader().font()
        # font.setBold(True)
        # self.tableWidget.horizontalHeader().setFont(font)
        # 标题字体加粗

        # tableWidget_2
        self.tableWidget_2.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        # 中表格铺满整个QTableWidget控件
        self.tableWidget_2.setEditTriggers(QAbstractItemView.NoEditTriggers)
        # 中表格不可编辑
        self.tableWidget_2.verticalHeader().setVisible(False)
        # 隐藏竖直滚动条
        self.tableWidget_2.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.tableWidget_2.setSelectionBehavior(QAbstractItemView.SelectRows)

        # tableWidget_5
        self.tableWidget_5.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # 中表格铺满整个QTableWidget控件
        self.tableWidget_5.setEditTriggers(QAbstractItemView.NoEditTriggers)
        # 中表格不可编辑
        self.tableWidget_5.verticalHeader().setVisible(False)  # 隐藏垂直表头
        self.tableWidget_5.horizontalHeader().setVisible(False)  # 隐藏水平表头
        # 隐藏竖直滚动条
        self.tableWidget_5.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.tableWidget_5.setSelectionBehavior(QAbstractItemView.SelectRows)

        # tableWidget_3
        self.tableWidget_4.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget_3.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # 右表格铺满整个QTableWidget控件
        self.tableWidget_3.setEditTriggers(QAbstractItemView.NoEditTriggers)
        # 右表格不可编辑
        self.tableWidget_3.verticalHeader().setVisible(False)
        self.tableWidget_4.verticalHeader().setVisible(False)  # 隐藏垂直表头
        self.tableWidget_4.horizontalHeader().setVisible(False)  # 隐藏水平表头
        # 隐藏竖直滚动条
        self.tableWidget_4.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.tableWidget_7.verticalHeader().setVisible(False)  # 隐藏垂直表头
        self.tableWidget_7.horizontalHeader().setVisible(False)  # 隐藏水平表头
        self.tableWidget_6.verticalHeader().setVisible(False)  # 隐藏垂直表头
        self.tableWidget_6.horizontalHeader().setVisible(False)  # 隐藏水平表头
        # self.tableWidget_8.verticalHeader().setVisible(False)  # 隐藏垂直表头
        self.tableWidget_8.horizontalHeader().setVisible(False)  # 隐藏水平表头
        self.tableWidget_block = ''
        globalData.scrollBar_A = self.tableWidget.verticalScrollBar()
        globalData.scrollBar_A.valueChanged.connect(self.verticalScrollBarChanged_a)
        globalData.scrollBar_B = self.tableWidget_2.verticalScrollBar()
        globalData.scrollBar_B.valueChanged.connect(self.verticalScrollBarChanged_a)
        globalData.scrollBar_C = self.tableWidget_3.verticalScrollBar()
        globalData.scrollBar_C.valueChanged.connect(self.verticalScrollBarChanged_a)

        globalData.scrollBar_A2 = self.tableWidget_4.verticalScrollBar()
        globalData.scrollBar_A2.valueChanged.connect(self.verticalScrollBarChanged_b)
        globalData.scrollBar_B2 = self.tableWidget_5.verticalScrollBar()
        globalData.scrollBar_B2.valueChanged.connect(self.verticalScrollBarChanged_b)
        globalData.scrollBar_C2 = self.tableWidget_6.verticalScrollBar()
        globalData.scrollBar_C2.valueChanged.connect(self.verticalScrollBarChanged_b)

    def verticalScrollBarChanged_a(self, e):
        globalData.scrollBar_A.setValue(e)
        globalData.scrollBar_B.setValue(e)
        globalData.scrollBar_C.setValue(e)

    def verticalScrollBarChanged_b(self, e):
        globalData.scrollBar_A2.setValue(e)
        globalData.scrollBar_B2.setValue(e)
        globalData.scrollBar_C2.setValue(e)

    def ImportVRLDA(self):
        if globalData.templateFlag:
            filePath1 = './assets/Template/VRLDA.csv'
        else:
            globalData.VRLDAisXLSX = False
            # 实例化QFileDialog
            dig = QFileDialog()
            # 设置可以打开任何文件
            dig.setFileMode(QFileDialog.AnyFile)
            # 文件过滤
            dig.setFilter(QDir.Files)

            if dig.exec_():
                # 接受选中文件的路径，默认为列表
                filenames = dig.selectedFiles()
                # 列表中的第一个元素即是文件路径，以只读的方式打开文件
                filePath1 = filenames[0]
        if is_used(filePath1):
            self.messageDialog('警告', '文件正在被占用')
        else:
            filename = filePath1.rsplit("/", 1)[1]
            if (re.search('.csv', filename)):
                globalData.VRLDApath = filePath1
                data = pd.read_csv(filePath1)
                data2 = np.array(data)
                data = np.array(data)[::-1]
                flag = 0
                self.tableWidget.setRowCount(0)
                self.tableWidget.clearContents()
                self.tableWidget_2.setRowCount(0)
                self.tableWidget_2.clearContents()
                self.tableWidget_3.setRowCount(0)
                self.tableWidget_3.clearContents()
                self.tableWidget_5.setRowCount(0)
                self.tableWidget_5.clearContents()
                self.tableWidget_4.setRowCount(0)
                self.tableWidget_4.clearContents()
                self.tableWidget_6.setRowCount(0)
                self.tableWidget_6.clearContents()
                self.tableWidget_7.setRowCount(0)
                self.tableWidget_7.clearContents()
                self.tableWidget_8.setRowCount(0)
                self.tableWidget_8.clearContents()
                globalData.HotSpots_N_List.clear()
                for i in range(len(data)):
                    item = data[i]
                    if (len(item) != 3 and flag == 0):
                        flag = 1
                        self.messageDialog('警告', '文件似乎不是VRLDA损伤数据，确实导入？')
                    globalData.HotSpots_N_List.append(data2[i][1])
                    self.tableWidget.insertRow(0)
                    for j in range(len(item)):
                        if j == 2:
                            # format(data[i][j], ".3f")
                            L1 = Decimal(data[i][j]).quantize(Decimal("0.000"))
                            z = str(L1)
                        else:
                            z = str(data[i][j])
                        # print(z)
                        item = QTableWidgetItem(z)
                        item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                        self.tableWidget.setItem(0, j, item)
                        if (i % 2 == 0):
                            self.tableWidget.item(0, j).setBackground(QBrush(QColor(244, 244, 244)))
            elif (re.search('.xlsx', filename)):
                # globalData.VRLDApath = filenames[0]
                workbook = xlrd.open_workbook(filePath1)
                sheet_names = workbook.sheet_names()
                if (len(sheet_names) == 1):
                    globalData.VRLDApath = copyXlsx(filePath1, globalData.cacheFilepath, 'VrldaCache.xlsx',
                                                    sheetName=None)
                else:
                    item, ok = QInputDialog.getItem(self, "请选择sheet", 'sheet列表', sheet_names, 0, False)
                    if ok and item:
                        sheetName = item
                    # print(sheetName)
                    globalData.VRLDApath = copyXlsx(filePath1, globalData.cacheFilepath, 'VrldaCache.xlsx',
                                                    sheetName)
                globalData.VRLDAisXLSX = True
                workbook = xlrd.open_workbook(globalData.VRLDApath)
                sheet_names = workbook.sheet_names()

                sheet = workbook.sheet_by_index(0)

                rows = sheet.nrows  # 获取有多少行
                cols = sheet.ncols  # 获取有多少列
                self.tableWidget.setRowCount(0)
                self.tableWidget.clearContents()
                self.tableWidget_2.setRowCount(0)
                self.tableWidget_2.clearContents()
                self.tableWidget_3.setRowCount(0)
                self.tableWidget_3.clearContents()
                self.tableWidget_5.setRowCount(0)
                self.tableWidget_5.clearContents()
                self.tableWidget_4.setRowCount(0)
                self.tableWidget_4.clearContents()
                self.tableWidget_6.setRowCount(0)
                self.tableWidget_6.clearContents()
                self.tableWidget_7.setRowCount(0)
                self.tableWidget_7.clearContents()
                self.tableWidget_8.setRowCount(0)
                self.tableWidget_8.clearContents()
                globalData.HotSpots_N_List.clear()
                if (cols != 3):
                    self.messageDialog('警告', '文件似乎不是VRLDA损伤数据，确实导入？')
                for i in range(1, rows):
                    self.tableWidget.insertRow(i - 1)
                    for j in range(0, cols):
                        # print(type(sheet.cell_value(i, j)))
                        value = sheet.cell_value(i, j)
                        # print(value)
                        if (i > 0 and j == 1):
                            globalData.HotSpots_N_List.append(value)
                            value = str(value).split('.')[0]
                            # print(globalData.HotSpots_N_List)
                        if j == 2 and i > 0:
                            # format(data[i][j], ".3f")
                            L1 = Decimal(float(value)).quantize(Decimal("0.000"))
                            z = str(L1)
                        else:
                            z = str(value)
                        # print(z)
                        item = QTableWidgetItem(z)
                        item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                        self.tableWidget.setItem(i - 1, j, item)
                        if (i - 1 % 2 == 0):
                            self.tableWidget.item(i - 1, j).setBackground(QBrush(QColor(244, 244, 244)))
            else:
                msg_box = QMessageBox(QMessageBox.Warning, '警告', '请选择csv或者xlsx文件')
                msg_box.exec_()

    def ImportLoadData(self):
        if globalData.templateFlag:
            filePath2 = './assets/Template/load.csv'
        else:

            dig = QFileDialog()
            dig.setFileMode(QFileDialog.AnyFile)
            dig.setFilter(QDir.Files)
            if dig.exec_():
                filenames = dig.selectedFiles()
                filePath2 = filenames[0]
        list = filePath2.rsplit("/", 1)
        filename = list[1]

        if is_used(filePath2):
            self.messageDialog('警告', '文件正在被占用')
        else:
            if re.search('.csv', filename):
                filePath = os.path.join(globalData.cacheFilepath, 'cache')
                if not os.path.exists(filePath):
                    os.mkdir(filePath)
                xlsxname = csv_to_xlsx(filePath, filePath2)
                globalData.LoadDatapath = os.path.join(globalData.cacheFilepath, 'cache', xlsxname).replace(
                    "\\",
                    '/')
                globalData.LoadDatapath_xlsx = copyXlsx(globalData.LoadDatapath, globalData.cacheFilepath,
                                                        'loadPathcache.xlsx', sheetName=None)
                BlockNameList = getblockNmaeListfromxlsx(globalData.LoadDatapath)
                globalData.bolckNameList = BlockNameList
                globalData.BlockCount = len(BlockNameList)
                self.tableWidget_2.setRowCount(0)
                self.tableWidget_2.clearContents()
                self.tableWidget_5.setRowCount(0)
                self.tableWidget_5.clearContents()
                self.tableWidget_7.setRowCount(0)
                self.tableWidget_7.clearContents()
                self.tableWidget_8.setRowCount(0)
                self.tableWidget_8.clearContents()
                self.tableWidget_2.setColumnCount(globalData.BlockCount)
                self.tableWidget_4.setColumnCount(3)
                self.tableWidget_5.setColumnCount(globalData.BlockCount)
                self.tableWidget_7.setColumnCount(globalData.BlockCount)
                ShowbolckNameList = []
                for i in globalData.bolckNameList:
                    ShowbolckNameList.append(i.split(' ')[0])
                self.tableWidget_2.setHorizontalHeaderLabels(ShowbolckNameList)
                self.tableWidget_2.resizeColumnsToContents()
            elif re.search('.xlsx', filename):
                globalData.loadDataisXLSX = True
                # globalData.LoadDatapath = copyXlsx(filenames[0], globalData.cacheFilepath, flag=1, self=MyMainForm)
                list = filePath2.rsplit("/", 1)
                workbook = xlrd.open_workbook(filePath2)
                sheet_names = workbook.sheet_names()
                sheetName = None
                if (len(sheet_names) == 1):
                    # 根据sheet索引或者名称获取sheet内容
                    sheet = workbook.sheet_by_index(0)  # sheet索引从0开始
                else:
                    item, ok = QInputDialog.getItem(self, "请选择sheet", 'sheet列表', sheet_names, 0, False)
                    if ok and item:
                        sheetName = item
                        sheet = workbook.sheet_by_name(item)
                rows = sheet.nrows  # 获取有多少行
                cols = sheet.ncols  # 获取有多少列
                # sheet.cell_value(第几行,第几列)
                datetime_object = datetime.datetime.now()

                time = str(datetime_object).split(' ')[0]
                xlsxname = 'Load_ IterationN' + time + '.xlsx'
                demoxlsxpath = os.path.join(globalData.cacheFilepath, 'cache', xlsxname)
                workbook2 = xlsxwriter.Workbook(demoxlsxpath)  # 创建一个excel文件
                # 在文件中创建一个名为TEST的sheet,不加名字默认为sheet1
                worksheet = workbook2.add_worksheet(u'Sheet1')
                for i in range(0, rows):
                    for j in range(0, cols):
                        value = sheet.cell_value(i, j)
                        worksheet.write(i, j, value)
                workbook2.close()
                globalData.LoadDatapath = demoxlsxpath.replace('\\', '/')
                # globalData.LoadDatapath = copyXlsx(filenames[0], globalData.cacheFilepath, flag=1, self=MyMainForm)
                # --------------------------------------
                globalData.LoadDatapath_xlsx = copyXlsx(filePath2, globalData.cacheFilepath,
                                                        'loadPathcache.xlsx', sheetName=sheetName)
                BlockNameList = getblockNmaeListfromxlsx(globalData.LoadDatapath)
                globalData.bolckNameList = BlockNameList
                globalData.BlockCount = len(BlockNameList)
                self.tableWidget_2.setRowCount(0)
                self.tableWidget_2.clearContents()
                self.tableWidget_5.setRowCount(0)
                self.tableWidget_5.clearContents()
                self.tableWidget_7.setRowCount(0)
                self.tableWidget_7.clearContents()
                self.tableWidget_8.setRowCount(0)
                self.tableWidget_8.clearContents()
                self.tableWidget_2.setColumnCount(globalData.BlockCount)
                self.tableWidget_4.setColumnCount(3)
                self.tableWidget_5.setColumnCount(globalData.BlockCount)
                self.tableWidget_7.setColumnCount(globalData.BlockCount)
                ShowbolckNameList = []
                for i in globalData.bolckNameList:
                    ShowbolckNameList.append(i.split(' ')[0])
                self.tableWidget_2.setHorizontalHeaderLabels(ShowbolckNameList)
                self.tableWidget_2.resizeColumnsToContents()
            else:
                msg_box = QMessageBox(QMessageBox.Warning, '警告', '请选择csv或者xlsx文件')
                msg_box.exec_()

    def ImportCYCLE(self):
        if len(globalData.bolckNameList) == 0:
            self.messageDialog('警告', '请先导入载荷数据')
        elif len(globalData.HotSpots_N_List) == 0:
            self.messageDialog('警告', '请先导入VRLDA数据')
        else:
            if globalData.templateFlag:
                filePath3 = './assets/Template/CYCLE.csv'

            else:
                # 实例化QFileDialog
                dig = QFileDialog()
                # 设置可以打开任何文件
                dig.setFileMode(QFileDialog.AnyFile)
                # 文件过滤
                dig.setFilter(QDir.Files)
                if dig.exec_():
                    # 接受选中文件的路径，默认为列表
                    filenames = dig.selectedFiles()
                    filePath3 = filenames[0]
            list = filePath3.rsplit("/", 1)
            filename = list[1]
            if is_used(filePath3):
                self.messageDialog('警告', '文件正在被占用')
            else:
                xlsxPath = ''
                if re.search('.csv', filename):
                    filepath = list[0]
                    xlsxname = csv_to_xlsx_pd(globalData.cacheFilepath, filepath, filename)
                    # csv转为xlsx
                    xlsxPath = os.path.join(globalData.cacheFilepath, 'cache', xlsxname).replace('\\', '/')
                elif re.search('.xlsx', filename):
                    xlsxPath = filePath3

                else:
                    msg_box = QMessageBox(QMessageBox.Warning, '警告', '请选择csv文件')
                    msg_box.exec_()
                workbook = xlrd.open_workbook(xlsxPath)
                sheet_names = workbook.sheet_names()
                if (len(sheet_names) == 1):
                    # 根据sheet索引或者名称获取sheet内容
                    sheet = workbook.sheet_by_index(0)  # sheet索引从0开始
                else:
                    item, ok = QInputDialog.getItem(self, "请选择sheet", 'sheet列表', sheet_names, 0, False)
                    if ok and item:
                        sheet = workbook.sheet_by_name(item)
                rows = sheet.nrows
                cols = sheet.ncols
                self.tableWidget_2.setRowCount(0)
                self.tableWidget_2.clearContents()
                self.tableWidget_3.setRowCount(0)
                self.tableWidget_3.clearContents()
                self.tableWidget_4.setRowCount(0)
                self.tableWidget_4.clearContents()
                self.tableWidget_5.setRowCount(0)
                self.tableWidget_5.clearContents()
                self.tableWidget_6.setRowCount(0)
                self.tableWidget_6.clearContents()
                self.tableWidget_7.setRowCount(0)
                self.tableWidget_7.clearContents()
                self.tableWidget_8.setRowCount(0)
                self.tableWidget_8.clearContents()
                self.tableWidget_8.setColumnCount(1)
                self.tableWidget_8.setVerticalHeaderLabels(['TOTAL = '])
                self.tableWidget_8.setEditTriggers(QAbstractItemView.NoEditTriggers)
                self.tableWidget_7.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                self.tableWidget_6.setColumnCount(1)
                self.tableWidget_6.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                self.tableWidget_8.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                self.tableWidget_6.setEditTriggers(QAbstractItemView.NoEditTriggers)
                self.tableWidget_7.cellChanged.connect(self.calculate)
                # 清除事件绑定
                globalData.firstFlag = 0

                blockNameofXLSX = []
                for l in range(1, cols):
                    s = sheet.cell_value(0, l)
                    blockNameofXLSX.append(s.split(' ')[0])
                a1 = np.array(blockNameofXLSX)
                # print(a1)
                x = []
                for i in globalData.bolckNameList:
                    x.append(i.split(' ')[0])

                a2 = np.array(x)
                # print(a2)
                ifContains = np.in1d(a1, a2)
                # print(ifContains)
                kif = False
                if kif in ifContains:
                    Nstr = ""
                    for i in range(len(ifContains)):
                        if not ifContains[i]:
                            Nstr += (a1[i] + ' ')
                    noticeStr = '可能存在表头不一致，或者Cycle表中的{}列在载荷数据中未定义，请检查载荷数据和Cycle损伤数据表头的一致性，重新导入'.format(Nstr)
                    self.messageDialog('提示', noticeStr)
                dataListfortableWidget_2 = []
                zeroList = []
                dataListfortableWidget_4 = []
                dataListfortableWidget_5 = []
                for i in range(len(globalData.bolckNameList) + 1):
                    zeroList.append(0)
                for item in globalData.HotSpots_N_List:
                    flag = 0
                    for r in range(1, rows):
                        if (item == sheet.cell_value(r, 0)):
                            ltemList = []
                            for block in globalData.bolckNameList:
                                flag2 = 0
                                for l in range(1, cols):
                                    s = sheet.cell_value(0, l)
                                    if (re.search(s, block)):
                                        value = sheet.cell_value(r, l)
                                        ltemList.append(ScientificEnumerationFormatting(value))
                                        flag2 = 1
                                        break
                                if (flag2 == 0):
                                    ltemList.append(0)
                            dataListfortableWidget_2.append(ltemList)
                            flag = 1
                            break
                    if (flag == 0):
                        dataListfortableWidget_2.append(zeroList)
                HotSpots_N_List0FXLSX = []
                for r in range(1, rows):
                    HotSpots_N_List0FXLSX.append(sheet.cell_value(r, 0))
                maskList = np.in1d(HotSpots_N_List0FXLSX, globalData.HotSpots_N_List)
                # print(maskList)
                for ind in range(len(maskList)):
                    if not maskList[ind]:
                        dataListfortableWidget_4.append(['ADD', sheet.cell_value(ind + 1, 0)])
                        ltemList = []
                        for block in globalData.bolckNameList:
                            flag2 = 0
                            for l in range(1, cols):
                                s = sheet.cell_value(0, l)
                                if (re.search(s, block)):
                                    value = sheet.cell_value(ind + 1, l)
                                    ltemList.append(ScientificEnumerationFormatting(value))
                                    flag2 = 1
                                    break
                            if (flag2 == 0):
                                ltemList.append(0)
                        dataListfortableWidget_5.append(ltemList)

                flagforwid_7 = 0
                # self.tableWidget_2.resizeColumnsToContents()
                colorIndex = 0
                for rowIndex in range(len(dataListfortableWidget_2)):
                    rowData = dataListfortableWidget_2[rowIndex]
                    if (flagforwid_7 == 0):
                        self.tableWidget_2.insertRow(rowIndex)
                        self.tableWidget_7.insertRow(flagforwid_7)
                        for j in range(len(rowData)):
                            item = QTableWidgetItem(str(dataListfortableWidget_2[rowIndex][j]))
                            item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                            self.tableWidget_2.setItem(rowIndex, j, item)
                            colorIndex += 1
                            if (rowIndex % 2 == 0 and self.tableWidget_2.item(rowIndex, j)):
                                self.tableWidget_2.item(rowIndex, j).setBackground(
                                    QBrush(QColor(244, 244, 244)))
                            item_7 = QTableWidgetItem(str(0))
                            item_7.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                            self.tableWidget_7.setItem(0, j, item_7)
                        flagforwid_7 = 1

                    else:
                        self.tableWidget_2.insertRow(rowIndex)
                        for j in range(len(rowData)):
                            item = QTableWidgetItem(str(dataListfortableWidget_2[rowIndex][j]))
                            item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                            self.tableWidget_2.setItem(rowIndex, j, item)
                            if (rowIndex % 2 == 0 and self.tableWidget_2.item(rowIndex, j)):
                                self.tableWidget_2.item(rowIndex, j).setBackground(
                                    QBrush(QColor(244, 244, 244)))

                for rowIndex in range(len(dataListfortableWidget_4)):
                    rowData = dataListfortableWidget_4[rowIndex]
                    self.tableWidget_4.insertRow(rowIndex)
                    for j in range(len(rowData)):
                        item = QTableWidgetItem(str(dataListfortableWidget_4[rowIndex][j]))
                        item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                        self.tableWidget_4.setItem(rowIndex, j, item)
                        if (rowIndex % 2 == 0 and self.tableWidget_4.item(rowIndex, j)):
                            self.tableWidget_4.item(rowIndex, j).setBackground(QBrush(QColor(244, 244, 244)))
                # print(dataListfortableWidget_5)
                for rowIndex in range(len(dataListfortableWidget_5)):
                    rowData = dataListfortableWidget_5[rowIndex]
                    self.tableWidget_5.insertRow(rowIndex)
                    self.tableWidget_6.insertRow(rowIndex)
                    for j in range(len(rowData)):
                        item = QTableWidgetItem(str(dataListfortableWidget_5[rowIndex][j]))
                        item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                        self.tableWidget_5.setItem(rowIndex, j, item)
                        if (rowIndex % 2 == 0 and self.tableWidget_5.item(rowIndex, j)):
                            self.tableWidget_5.item(rowIndex, j).setBackground(QBrush(QColor(244, 244, 244)))
                self.tableWidget_7.cellChanged.connect(self.calculate)
                globalData.cycleFlag = 1
                self.calculate(0, 0)

            # if re.search('.csv', filename):
            #     if is_used(filenames[0]):
            #         self.messageDialog('警告', '文件正在被占用')
            #     else:
            #         filepath = list[0]
            #         xlsxname = csv_to_xlsx_pd(globalData.cacheFilepath, filepath, filename)
            #         # csv转为xlsx
            #         xlsxPath = os.path.join(globalData.cacheFilepath, 'cache', xlsxname).replace('\\', '/')
            #
            # elif re.search('.xlsx', filename):
            #     if is_used(filenames[0]):
            #         self.messageDialog('警告', '文件正在被占用')
            #     else:
            #         xlsxPath=filenames[0]
            #
            # else:
            #     msg_box = QMessageBox(QMessageBox.Warning, '警告', '请选择csv文件')
            #     msg_box.exec_()

    def calculate(self, row, col):
        if (not self.tableWidget_7.item(row, col).text().isdigit()):
            self.messageDialog('警告', '只能输入数字')
            item_7 = QTableWidgetItem(str(0))
            item_7.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
            self.tableWidget_7.setItem(row, col, item_7)
        else:
            ratioList = []
            tableWidget_6_ratioList = []
            tableWidget_7_List = []
            cyclesum = 0
            for i in range(globalData.BlockCount):
                if self.tableWidget_7.item(0, i):
                    cyclesum += int(self.tableWidget_7.item(0, i).text())
                    tableWidget_7_List.append(int(self.tableWidget_7.item(0, i).text()))
                else:
                    cyclesum += int(0)
                    tableWidget_7_List.append(int(0))

            for rowIndex in range(len(globalData.HotSpots_N_List)):
                sumRadio = 0
                for index in range(globalData.BlockCount):
                    y = int(tableWidget_7_List[index])
                    if y != 0:
                        if self.tableWidget_2.item(rowIndex, index):
                            a = self.tableWidget_2.item(rowIndex, index).text()
                            if (a != '0'):
                                x = ScientificEnumeration2Number(a)
                            else:
                                x = float(0)
                            t = x * y
                            sumRadio += t
                ratioList.append(f"{sumRadio:.2e}")
            globalDataRatioList = []
            for i in range(len(ratioList)):
                globalDataRatioList.append(ratioList[i])
            lowRadio = 0
            if globalDataRatioList:
                globalDataRatioList.sort(reverse=True)
                if len(globalDataRatioList) >= 3:
                    lowRadio = ScientificEnumeration2Number(globalDataRatioList[2])
                else:
                    lowRadio = ScientificEnumeration2Number(globalDataRatioList[len(globalDataRatioList) - 1])
            # print(globalDataRatioList)

            rows = self.tableWidget_5.rowCount()
            for row in range(rows):
                sum = 0
                for index in range(globalData.BlockCount):
                    y = int(tableWidget_7_List[index])
                    if y != 0:
                        if self.tableWidget_5.item(row, index):
                            a = self.tableWidget_5.item(row, index).text()
                            if (a != '0'):
                                x = ScientificEnumeration2Number(a)
                            else:
                                x = float(0)
                            t = x * y
                            sum += t
                tableWidget_6_ratioList.append(f"{sum:.2e}")

            for i in range(len(tableWidget_6_ratioList)):
                if int(ScientificEnumeration2Number(tableWidget_6_ratioList[i])) < lowRadio:
                    tableWidget_6_ratioList[i] = 'OK'
                else:
                    tableWidget_6_ratioList[i] = 'Nok'

            if (globalData.firstFlag == 0):
                # print('1-------', ratioList)
                for i in range(len(ratioList)):
                    self.tableWidget_3.insertRow(i)
                    item = QTableWidgetItem(str(ratioList[i]))
                    item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                    self.tableWidget_3.setItem(i, 0, item)
                    if (i % 2 == 0 and self.tableWidget_3.item(i, 0)):
                        self.tableWidget_3.item(i, 0).setBackground(QBrush(QColor(244, 244, 244)))
                for i in range(len(tableWidget_6_ratioList)):
                    item = QTableWidgetItem(str(tableWidget_6_ratioList[i]))
                    item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                    self.tableWidget_6.setItem(i, 0, item)
                    if (i % 2 == 0 and self.tableWidget_6.item(i, 0)):
                        self.tableWidget_6.item(i, 0).setBackground(QBrush(QColor(244, 244, 244)))
                globalData.firstFlag = 1
            else:
                # print('2-------', ratioList)
                for i in range(len(ratioList)):
                    item = QTableWidgetItem(str(ratioList[i]))
                    item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                    self.tableWidget_3.setItem(i, 0, item)
                for i in range(len(tableWidget_6_ratioList)):
                    item = QTableWidgetItem(str(tableWidget_6_ratioList[i]))
                    item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                    self.tableWidget_6.setItem(i, 0, item)
            if (self.tableWidget_8.rowCount() == 0):
                self.tableWidget_8.insertRow(0)
            self.tableWidget_8.setItem(0, 0, QTableWidgetItem(str(cyclesum)))
            self.tableWidget_8.setVerticalHeaderLabels(['TOTAL = '])

    def HorSectionClicked(self, index):
        if len(globalData.HotSpots_N_List) != 0:
            globalData.blockIndex = index
            blockName = globalData.bolckNameList[index]
            if not globalData.LoadDatapath_xlsx:
                self.messageDialog('警告', '请先导入文件数据')
            else:
                ans = getxlsxData(globalData.LoadDatapath_xlsx, blockName, globalData.GVWindexList)
                print(ans)
                xlsxPath = ans[2]
                globalData.left = ans[0]
                globalData.right = ans[1]
                globalData.GVWindexList = ans[3]
                # print(globalData.GVWindexList)
                workbook = xlrd.open_workbook(xlsxPath)
                sheet = workbook.sheet_by_index(0)
                rows = sheet.nrows
                cols = sheet.ncols
                globalData.dialog = ChildWin_block()
                self.tableWidget_block = globalData.dialog.tableWidget
                globalData.dialog.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                globalData.dialog.tableWidget.setColumnCount(cols)
                globalData.dialog.tableWidget.horizontalHeader().setVisible(False)  # 隐藏水平表头
                # print(sheet.row(1))
                titleList = sheet.row(1)
                if (str(sheet.cell_value(1, 4)) == str(sheet.cell_value(1, len(titleList) - 1))):
                    globalData.GVWflag = len(titleList) - 1
                # print(globalData.GVWflag)
                for i in range(rows):
                    row = globalData.dialog.tableWidget.rowCount()
                    globalData.dialog.tableWidget.insertRow(row)
                    for j in range(cols):
                        item = QTableWidgetItem(str(sheet.cell_value(i, j)))
                        if (i == 0 or i == 1 or j == 0 or j == 1 or j == 2 or j == 3):
                            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                        #     部分可编辑
                        item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                        globalData.dialog.tableWidget.setItem(row, j, item)
                globalData.dialog.tableWidget.setSpan(0, 4, 1, cols)
                globalData.dialog.tableWidget.cellChanged.connect(self.updateData)
                # globalData.dialog.tableWidget.cellEntered.connect(self.updateData)
                globalData.dialog.show()

                def pB_OK():
                    globalData.dialog.close()
                    os.remove(xlsxPath)

                globalData.dialog.buttonBox.clicked.connect(pB_OK)
                globalData.dialog.exec_()

        else:
            self.messageDialog('警告', '请先导入文件数据')

    def changeData(self, row, col):
        if globalData.updateFlag:
            globalData.updateFlag = False
            self.updateData(row, col)

    def updateData(self, row, col):
        if is_used(globalData.LoadDatapath_xlsx):
            self.messageDialog('警告', '文件正在被占用')
        else:
            globalData.editChangeFlag = 1
            value = self.tableWidget_block.item(row, col).text()
            if value:
                wb = load_workbook(filename=globalData.LoadDatapath_xlsx)
                ws = wb['Sheet1']
                if globalData.GVWflag != 0:
                    if col == 4 or col == globalData.GVWflag:
                        for GVWindex in globalData.GVWindexList:
                            if GVWindex < 26:
                                x = chr(int(GVWindex) + 65)
                                item = x + str(row + 1)
                                ws[item] = value
                            else:
                                count = int(GVWindex / 26)
                                t = int(GVWindex) - 26 * count
                                x = chr(t + 65)
                                x = chr(65 + count - 1) + x
                                item = x + str(row + 1)
                                print(item)
                                ws[item] = value
                            # if GVWindex >= 26 and GVWindex < 52:
                            #     t = int(GVWindex) - 26
                            #     x = chr(t + 65)
                            #     x = 'A' + x
                            #     item = x + str(row + 1)
                            #     ws[item] = value
                            # elif GVWindex >= 52 and GVWindex < 78:
                            #     t = int(GVWindex) - 52
                            #     x = chr(t + 65)
                            #     x = 'B' + x
                            #     item = x + str(row + 1)
                            #     ws[item] = value
                            # elif GVWindex >= 78 and GVWindex < 104:
                            #     t = int(GVWindex) - 78
                            #     x = chr(t + 65)
                            #     x = 'C' + x
                            #     item = x + str(row + 1)
                            #     ws[item] = value
                            # elif GVWindex >= 104 and GVWindex < 130:
                            #     t = int(GVWindex) - 104
                            #     x = chr(t + 65)
                            #     x = 'D' + x
                            #     item = x + str(row + 1)
                            #     ws[item] = value
                            # else:
                            #     self.messageDialog('警告', '数据量过大')
                if int(globalData.left) + col - 4 < 26:
                    x = chr(int(globalData.left) + col - 4 + 65)
                    item = x + str(row + 1)
                    ws[item] = value
                # elif int(globalData.left) + col - 4 >= 26 and int(globalData.left) + col - 4 < 26*2:
                #     t = int(globalData.left) - 26
                #     x = chr(t + col - 4 + 65)
                #     x = 'A' + x
                #     item = x + str(row + 1)
                #     ws[item] = value
                # elif int(globalData.left) + col - 4 >= 26*2 and int(globalData.left) + col - 4 < 26*3:
                #     t = int(globalData.left) - 26*2
                #     x = chr(t + col - 4 + 65)
                #     x = 'B' + x
                #     item = x + str(row + 1)
                #     ws[item] = value
                # elif int(globalData.left) + col - 4 >= 26*3 and int(globalData.left) + col - 4 < 26*4:
                #     t = int(globalData.left) - 26*3
                #     x = chr(t + col - 4 + 65)
                #     x = 'A' + x
                #     item = x + str(row + 1)
                #     ws[item] = value
                else:
                    count = int((int(globalData.left) + col - 4) / 26)
                    t = int(globalData.left) - 26 * count
                    x = chr(t + col - 4 + 65)
                    x = chr(65 + count - 1) + x
                    item = x + str(row + 1)
                    print(item)
                    ws[item] = value
                wb.save(globalData.LoadDatapath_xlsx)
            else:
                pass
        globalData.dialog.close()
        globalData.updateFlag = True
        self.HorSectionClicked(globalData.blockIndex)

    def noticeSave(self):
        if (globalData.editChangeFlag == 0):
            self.messageDialog('提示', '载荷数据没有做修改！')
        else:
            directory = QtWidgets.QFileDialog.getExistingDirectory(None, "选取文件夹", globalData.userFilepath)  # 起始路径
            if is_used(globalData.LoadDatapath):
                self.messageDialog('警告', '文件正在被占用')
            else:
                addData2xlsx(globalData.LoadDatapath_xlsx, globalData.LoadDatapath)
                washXlsx(globalData.LoadDatapath)
                filename = globalData.LoadDatapath.rsplit("/", 1)[1]
                newFilepath = os.path.join(directory, filename).replace('\\', '/')
                if os.path.exists(newFilepath):
                    if is_used(newFilepath):
                        self.messageDialog('警告', '文件正在被占用')
                    else:
                        shutil.copy(globalData.LoadDatapath, directory)
                        if os.path.isfile(newFilepath):
                            self.messageDialog('提示', '保存成功！')
                else:
                    shutil.copy(globalData.LoadDatapath, directory)
                    if os.path.isfile(newFilepath):
                        self.messageDialog('提示', '保存成功！')
                # print(newFilepath)
                # for i in range(3):
                #     if os.path.isfile(newFilepath):
                #         self.messageDialog('提示', '保存成功！')
                #         break
                #     else:
                #         time.sleep(2)
                #     if i == 3:
                #         self.messageDialog('提示', '保存失败，请重试')

    def helpMessage(self):
        mess = getHelp()
        self.messageDialog('提示', mess)

    def writeXlsx(self):
        if not globalData.writeXlsxPath:
            directory = QtWidgets.QFileDialog.getExistingDirectory(None, "选取文件夹", globalData.userFilepath)  # 起始路径
            globalData.writeXlsxPath = directory
        else:
            directory = globalData.writeXlsxPath
        if directory:
            if not globalData.Result_IterationName:
                datetime_object = datetime.datetime.now()
                time = str(datetime_object).split(' ')[0] + '_' + (str(datetime_object).split(' ')[1]).split('.')[
                    0].replace(':', '_')
                filename = 'Result_IterationN_' + time + '.xlsx'
                globalData.Result_IterationName = filename
            else:
                filename = globalData.Result_IterationName
            newFilepath = os.path.join(directory, filename).replace('\\', '/')
            if os.path.exists(newFilepath):
                if is_used(newFilepath):
                    self.messageDialog('警告', '文件正在被占用')
                else:
                    w = xlrd.open_workbook(newFilepath)
                    sheetLength = len(w.sheet_names()) + 1
                    # print(sheetLength)
                    sheetName = 'Sheet' + str(sheetLength)
                    wb = openpyxl.load_workbook(newFilepath)
                    wb.create_sheet(title=sheetName, index=sheetLength)
                    wb.save(newFilepath)
                    wb = load_workbook(filename=newFilepath)
                    fill = PatternFill("solid", fgColor="ffff00")
                    fill2 = PatternFill("solid", fgColor="ff0000")
                    ws = wb[sheetName]
                    ws['A1'] = 'HotSpots_Position'
                    ws['A1'].fill = fill
                    ws['B1'] = 'HotSpots_N'
                    ws['B1'].fill = fill
                    ws['C1'] = 'VRLDA_Damage'
                    ws['C1'].fill = fill

                    flag = 0
                    for i in range(len(globalData.bolckNameList)):
                        if i + 3 < 26:
                            x = chr(i + 65 + 3)
                        # elif i + 3 >= 26:
                        #     t = i + 3 - 26
                        #     x = chr(t + 65)
                        #     x = 'A' + x
                        else:
                            count = int((i + 3) / 26)
                            t = (i + 3) - 26 * count
                            x = chr(t + 65)
                            x = chr(65 + count - 1) + x
                        item = x + str(1)
                        # print('item----', item)
                        ws[item] = globalData.bolckNameList[i]
                        ws[item].fill = fill
                    if flag == 1:
                        self.messageDialog('警告', '数据过长')
                    if len(globalData.bolckNameList) + 3 < 26:
                        x = chr(len(globalData.bolckNameList) + 65 + 3)
                    else:
                        count = int((len(globalData.bolckNameList) + 3) / 26)
                        t = len(globalData.bolckNameList) + 3 - 26 * count
                        x = chr(t + 65)
                        x = chr(65 + count - 1) + x
                    # if len(globalData.bolckNameList) + 3 >= 26:
                    #     t = len(globalData.bolckNameList) + 3 - 26
                    #     x = 'A' + chr(t + 65)
                    # else:
                    #     x = chr(len(globalData.bolckNameList) + 65 + 3)
                    item = x + str(1)
                    ws[item] = 'RADIO'
                    ws[item].fill = fill

                    rows_1 = self.tableWidget.rowCount()
                    cols_1 = self.tableWidget.columnCount()
                    for rowindex in range(rows_1):
                        for colindex in range(cols_1):
                            item = chr(colindex + 65) + str(rowindex + 2)
                            ws[item] = self.tableWidget.item(rowindex, colindex).text() if self.tableWidget.item(
                                rowindex, colindex) else ''

                    rows_2 = self.tableWidget_2.rowCount()
                    cols_2 = self.tableWidget_2.columnCount()
                    for rowindex in range(rows_2):
                        for colindex in range(cols_2):
                            if colindex + 3 < 26:
                                x = chr(colindex + 65 + 3)
                            else:
                                count = int((colindex + 3) / 26)
                                t = colindex + 3 - 26 * count
                                x = chr(t + 65)
                                x = chr(65 + count - 1) + x
                            # if colindex + 3 >= 26:
                            #     t = colindex + 3 - 26
                            #     x = chr(t + 65)
                            #     x = 'A' + x
                            item = x + str(rowindex + 2)
                            ws[item] = self.tableWidget_2.item(rowindex, colindex).text() if self.tableWidget_2.item(
                                rowindex, colindex) else ''

                    rows_3 = self.tableWidget_3.rowCount()
                    cols_3 = self.tableWidget_3.columnCount()
                    for rowindex in range(rows_3):
                        for colindex in range(cols_3):
                            if colindex + len(globalData.bolckNameList) + 3 < 26:
                                x = chr(colindex + len(globalData.bolckNameList) + 65 + 3)
                            else:
                                count = int((colindex + len(globalData.bolckNameList) + 3) / 26)
                                t = colindex + len(globalData.bolckNameList) + 3 - 26 * count
                                x = chr(t + 65)
                                x = chr(65 + count - 1) + x
                            # if colindex + len(globalData.bolckNameList) + 3 >= 26:
                            #     t = colindex + len(globalData.bolckNameList) + 3 - 26
                            #     x = chr(t + 65)
                            #     x = 'A' + x
                            item = x + str(rowindex + 2)
                            ws[item] = self.tableWidget_3.item(rowindex, colindex).text() if self.tableWidget_3.item(
                                rowindex, colindex) else ''
                            ws[item].fill = fill2

                    interval = 2

                    rows_4 = self.tableWidget_4.rowCount()
                    cols_4 = self.tableWidget_4.columnCount()
                    for rowindex in range(rows_4):
                        for colindex in range(cols_4):
                            item = chr(colindex + 65) + str(rowindex + rows_1 + 2 + interval)
                            ws[item] = self.tableWidget_4.item(rowindex, colindex).text() if self.tableWidget_4.item(
                                rowindex, colindex) else ''

                    rows_5 = self.tableWidget_5.rowCount()
                    cols_5 = self.tableWidget_5.columnCount()
                    for rowindex in range(rows_5):
                        for colindex in range(cols_5):
                            if colindex + 3 < 26:
                                x = chr(colindex + 65 + 3)
                            else:
                                count = int((colindex + 3) / 26)
                                t = colindex + 3 - 26 * count
                                x = chr(t + 65)
                                x = chr(65 + count - 1) + x
                            # if colindex + 3 >= 26:
                            #     t = colindex + 3 - 26
                            #     x = chr(t + 65)
                            #     x = 'A' + x
                            item = x + str(rowindex + rows_1 + 2 + interval)
                            ws[item] = self.tableWidget_5.item(rowindex, colindex).text() if self.tableWidget_5.item(
                                rowindex, colindex) else ''

                    rows_6 = self.tableWidget_6.rowCount()
                    cols_6 = self.tableWidget_6.columnCount()
                    for rowindex in range(rows_6):
                        for colindex in range(cols_6):
                            if colindex + len(globalData.bolckNameList) + 3 < 26:
                                x = chr(colindex + len(globalData.bolckNameList) + 65 + 3)
                            else:
                                count = int((colindex + len(globalData.bolckNameList) + 3) / 26)
                                t = colindex + len(globalData.bolckNameList) + 3 - 26 * count
                                x = chr(t + 65)
                                x = chr(65 + count - 1) + x
                            # if colindex + len(globalData.bolckNameList) + 3 >= 26:
                            #     t = colindex + len(globalData.bolckNameList) + 3 - 26
                            #     x = chr(t + 65)
                            #     x = 'A' + x
                            item = x + str(rowindex + rows_1 + 2 + interval)
                            ws[item] = self.tableWidget_6.item(rowindex, colindex).text() if self.tableWidget_6.item(
                                rowindex, colindex) else ''
                            ws[item].fill = fill2
                    rows_7 = self.tableWidget_7.rowCount()
                    cols_7 = self.tableWidget_7.columnCount()
                    for rowindex in range(rows_7):
                        for colindex in range(cols_7):
                            if colindex + 3 < 26:
                                x = chr(colindex + 65 + 3)
                            else:
                                count = int((colindex + 3) / 26)
                                t = colindex + 3 - 26 * count
                                x = chr(t + 65)
                                x = chr(65 + count - 1) + x
                            # if colindex + 3 >= 26:
                            #     t = colindex + 3 - 26
                            #     x = chr(t + 65)
                            #     x = 'A' + x
                            item = x + str(rowindex + rows_1 + rows_4 + 2 + interval * 2)
                            ws[item] = self.tableWidget_7.item(rowindex, colindex).text() if self.tableWidget_7.item(
                                rowindex, colindex) else ''

                    rows_8 = self.tableWidget_8.rowCount()
                    cols_8 = self.tableWidget_8.columnCount()
                    for rowindex in range(rows_8):
                        for colindex in range(cols_8):
                            if colindex + len(globalData.bolckNameList) + 3 < 26:
                                x = chr(colindex + len(globalData.bolckNameList) + 65 + 3)
                            else:
                                count = int((colindex + len(globalData.bolckNameList) + 3) / 26)
                                t = colindex + len(globalData.bolckNameList) + 3 - 26 * count
                                x = chr(t + 65)
                                x = chr(65 + count - 1) + x
                            # if colindex + len(globalData.bolckNameList) + 3 >= 26:
                            #     t = colindex + len(globalData.bolckNameList) + 3 - 26
                            #     x = chr(t + 65)
                            #     x = 'A' + x
                            item = x + str(rowindex + rows_1 + rows_4 + 2 + interval * 2)
                            ws[item] = self.tableWidget_8.item(rowindex, colindex).text() if self.tableWidget_8.item(
                                rowindex, colindex) else ''
                            ws[item].fill = fill2

                    wb.save(newFilepath)
                    if os.path.exists(newFilepath):
                        self.messageDialog('提示', '保存成功')

            else:
                sheetName = 'Sheet1'
                workbook = xlsxwriter.Workbook(newFilepath)
                style1 = workbook.add_format({
                    "fg_color": "yellow",  # 单元格的背景颜色
                    "bold": 1,  # 字体加粗
                    "align": "center",  # 对齐方式
                    "valign": "vcenter",  # 字体对齐方式
                    "font_color": "red"  # 字体颜色

                })
                style2 = workbook.add_format({
                    "fg_color": "red",  # 单元格的背景颜色
                    "align": "center",  # 对齐方式
                    "valign": "vcenter",  # 字体对齐方式

                })
                worksheet = workbook.add_worksheet(sheetName)
                worksheet.write(0, 0, 'HotSpots_Position', style1)
                worksheet.write(0, 1, 'HotSpots_N', style1)
                worksheet.write(0, 2, 'VRLDA_Damage', style1)
                for i in range(len(globalData.bolckNameList)):
                    worksheet.write(0, i + 3, globalData.bolckNameList[i], style1)
                worksheet.write(0, len(globalData.bolckNameList) + 3, 'RADIO', style1)
                rows_1 = self.tableWidget.rowCount()
                cols_1 = self.tableWidget.columnCount()
                for rowindex in range(rows_1):
                    for colindex in range(cols_1):
                        worksheet.write(rowindex + 1, colindex,
                                        self.tableWidget.item(rowindex, colindex).text() if self.tableWidget.item(
                                            rowindex, colindex) else '')
                rows_2 = self.tableWidget_2.rowCount()
                cols_2 = self.tableWidget_2.columnCount()
                for rowindex in range(rows_2):
                    for colindex in range(cols_2):
                        worksheet.write(rowindex + 1, colindex + 3,
                                        self.tableWidget_2.item(rowindex, colindex).text() if self.tableWidget_2.item(
                                            rowindex, colindex) else '')

                rows_3 = self.tableWidget_3.rowCount()
                cols_3 = self.tableWidget_3.columnCount()
                for rowindex in range(rows_3):
                    for colindex in range(cols_3):
                        worksheet.write(rowindex + 1, colindex + 3 + len(globalData.bolckNameList),
                                        self.tableWidget_3.item(rowindex, colindex).text() if self.tableWidget_3.item(
                                            rowindex, colindex) else '', style2)

                interval = 2
                rows_4 = self.tableWidget_4.rowCount()
                cols_4 = self.tableWidget_4.columnCount()
                for rowindex in range(rows_4):
                    for colindex in range(cols_4):
                        worksheet.write(rowindex + 1 + rows_1 + interval, colindex,
                                        self.tableWidget_4.item(rowindex, colindex).text() if self.tableWidget_4.item(
                                            rowindex, colindex) else '')

                rows_5 = self.tableWidget_5.rowCount()
                cols_5 = self.tableWidget_5.columnCount()
                for rowindex in range(rows_5):
                    for colindex in range(cols_5):
                        worksheet.write(rowindex + 1 + rows_1 + interval, colindex + 3,
                                        self.tableWidget_5.item(rowindex, colindex).text() if self.tableWidget_5.item(
                                            rowindex, colindex) else '')

                rows_6 = self.tableWidget_6.rowCount()
                cols_6 = self.tableWidget_6.columnCount()
                for rowindex in range(rows_6):
                    for colindex in range(cols_6):
                        worksheet.write(rowindex + 1 + rows_1 + interval, colindex + 3 + len(globalData.bolckNameList),
                                        self.tableWidget_6.item(rowindex, colindex).text() if self.tableWidget_6.item(
                                            rowindex, colindex) else '', style2)

                rows_7 = self.tableWidget_7.rowCount()
                cols_7 = self.tableWidget_7.columnCount()
                for rowindex in range(rows_7):
                    for colindex in range(cols_7):
                        worksheet.write(rowindex + 1 + rows_1 + rows_4 + interval * 2, colindex + 3,
                                        self.tableWidget_7.item(rowindex, colindex).text() if self.tableWidget_7.item(
                                            rowindex, colindex) else '')

                rows_8 = self.tableWidget_8.rowCount()
                cols_8 = self.tableWidget_8.columnCount()
                for rowindex in range(rows_8):
                    for colindex in range(cols_8):
                        worksheet.write(rowindex + 1 + rows_1 + rows_4 + interval * 2,
                                        colindex + 3 + len(globalData.bolckNameList),
                                        self.tableWidget_8.item(rowindex, colindex).text() if self.tableWidget_8.item(
                                            rowindex, colindex) else '')
                workbook.close()
                if os.path.exists(newFilepath):
                    self.messageDialog('提示', '保存成功')

    def writetoCSV(self):
        if (globalData.cycleFlag == 0):
            self.messageDialog('提示', '无数据！')
        else:
            file1 = globalData.VRLDApath
            list = file1.rsplit("/", 1)
            filename = list[1]
            filepath = list[0]
            if globalData.VRLDAisXLSX:
                ansXlsxpath = globalData.VRLDApath
            else:
                xlsxname = csv_to_xlsx(os.path.join(globalData.cacheFilepath, 'cache'), file1)
                # xlsxname = csv_to_xlsx_pd(globalData.cacheFilepath, filepath, filename)
                ansXlsxpath = os.path.join(globalData.cacheFilepath, 'cache', xlsxname).replace('\\', '/')
            wb = load_workbook(filename=ansXlsxpath)
            fill = PatternFill("solid", fgColor="8064a2")
            fill2 = PatternFill("solid", fgColor="ffeb9c")
            ws = wb['Sheet1']
            for i in range(len(globalData.bolckNameList)):
                x = chr(i + 68)
                item = x + str(1)
                ws[item] = globalData.bolckNameList[i]
                ws[item].fill = fill
            ws[chr(len(globalData.bolckNameList) + 68) + str(1)] = 'RADIO'
            ws[chr(len(globalData.bolckNameList) + 68) + str(1)].fill = fill2
            rows_2 = self.tableWidget_2.rowCount()
            cols_2 = self.tableWidget_2.columnCount()
            for rowindex in range(rows_2):
                for colindex in range(cols_2):
                    item = chr(colindex + 68) + str(rowindex + 2)
                    ws[item] = self.tableWidget_2.item(rowindex, colindex).text()
                    if (colindex == cols_2 - 1):
                        item_3 = chr(cols_2 + 68) + str(rowindex + 2)
                        ws[item_3] = self.tableWidget_3.item(rowindex, 0).text()
                        ws[item_3].fill = fill2
            interval = 4

            # tableWidget_4
            rows_4 = self.tableWidget_4.rowCount()
            cols_4 = self.tableWidget_4.columnCount()
            for rowindex in range(rows_4):
                for colindex in range(cols_4):
                    item = chr(colindex + 65) + str(rowindex + interval + rows_2)
                    if self.tableWidget_4.item(rowindex, colindex):
                        x = self.tableWidget_4.item(rowindex, colindex).text()
                        ws[item] = x
                # if (colindex == cols_4 - 1):
                #     item_3 = chr(cols_4 + 65) + str(rowindex + interval + rows_2)
                #     ws[item_3] = self.tableWidget_6.item(rowindex, 0).text()
                #     ws[item_3].fill = fill2

            # tableWidget_5
            rows_5 = self.tableWidget_5.rowCount()
            cols_5 = self.tableWidget_5.columnCount()
            for rowindex in range(rows_5):
                for colindex in range(cols_5):
                    item = chr(colindex + 68) + str(rowindex + interval + rows_2)
                    ws[item] = self.tableWidget_5.item(rowindex, colindex).text()
                    if (colindex == cols_5 - 1):
                        item_3 = chr(cols_5 + 68) + str(rowindex + interval + rows_2)
                        ws[item_3] = self.tableWidget_6.item(rowindex, 0).text()
                        ws[item_3].fill = fill2
            # tableWidget_7
            cols_7 = self.tableWidget_7.columnCount()

            for colindex in range(cols_7):
                item = chr(colindex + 68) + str(interval * 2 + rows_2 + rows_5)
                ws[item] = self.tableWidget_7.item(0, colindex).text()
                if (colindex == cols_7 - 1):
                    item_3 = chr(cols_7 + 68) + str(interval * 2 + rows_2 + rows_5)
                    ws[item_3] = self.tableWidget_8.item(0, 0).text()
                    ws[item_3].fill = fill2
            wb.save(ansXlsxpath)

            directory = QtWidgets.QFileDialog.getExistingDirectory(None, "选取文件夹", globalData.userFilepath)  # 起始路径
            if directory:
                datetime_object = datetime.datetime.now()
                time = str(datetime_object).split(' ')[0]
                list = ansXlsxpath.rsplit("/", 1)
                filename = 'Result_IterationN_' + time + '.xlsx'
                filepath = list[0]
                newFilepath = os.path.join(directory, filename).replace('\\', '/')
                # saveFile = os.path.join(filepath, filename).replace('\\', '/')
                if os.path.exists(newFilepath):
                    os.remove(newFilepath)
                shutil.copy(ansXlsxpath, directory)
                yuanname = os.path.join(directory, list[1]).replace('\\', '/')
                os.rename(yuanname, newFilepath)
                for i in range(3):
                    if os.path.isfile(newFilepath):
                        self.messageDialog('提示', '保存成功！')
                        break
                    else:
                        time.sleep(2)
                    if i == 2:
                        self.messageDialog('提示', '保存失败，请重试')

    def messageDialog(self, type, message):
        # 核心功能代码就两行，可以加到需要的地方
        if type == '警告':
            x = QMessageBox.Warning
        if type == '提示':
            x = QMessageBox.Information
        msg_box = QMessageBox(x, type, message)
        msg_box.exec_()

    def clearAll(self):
        self.tableWidget.setRowCount(0)
        self.tableWidget.clearContents()
        self.tableWidget_2.setRowCount(0)
        self.tableWidget_2.clearContents()
        self.tableWidget_3.setRowCount(0)
        self.tableWidget_3.clearContents()
        self.tableWidget_5.setRowCount(0)
        self.tableWidget_5.clearContents()
        self.tableWidget_4.setRowCount(0)
        self.tableWidget_4.clearContents()
        self.tableWidget_6.setRowCount(0)
        self.tableWidget_6.clearContents()
        self.tableWidget_7.setRowCount(0)
        self.tableWidget_7.clearContents()
        self.tableWidget_8.setRowCount(0)
        self.tableWidget_8.clearContents()
        globalData.HotSpots_N_List.clear()
        ShowbolckNameList = []
        for i in globalData.bolckNameList:
            ShowbolckNameList.append('')
        self.tableWidget_2.setHorizontalHeaderLabels(ShowbolckNameList)
        globalData.HotSpots_N_List = []
        globalData.LoadDatapath = ''
        globalData.LoadDatapath_xlsx = ''
        globalData.VRLDAisXLSX = False
        globalData.writeXlsxPath = ''
        globalData.Result_IterationName = ''
        globalData.GVWindexList = []
        path = os.path.join(globalData.cacheFilepath, 'cache').replace('\\', '/')
        shutil.rmtree(path)
        os.mkdir(path)

    def importTemp(self):
        globalData.templateFlag = True
        self.ImportVRLDA()
        self.ImportLoadData()
        self.ImportCYCLE()
        globalData.templateFlag=False
    # if __name__ == "__main__":


def main():
    createLog()
    app = QApplication(sys.argv)
    # extra = {
    #
    #     # Button colors
    #     'danger': '#dc3545',
    #     'warning': '#ffc107',
    #     'success': '#17a2b8',
    #
    #     # Font
    #     'font-family': 'Roboto',
    # }
    # themeList = ['dark_amber.xml',
    #              'dark_blue.xml',
    #              'dark_cyan.xml',
    #              'dark_lightgreen.xml',
    #              'dark_pink.xml',
    #              'dark_purple.xml',
    #              'dark_red.xml',
    #              'dark_teal.xml',
    #              'dark_yellow.xml',
    #              'light_amber.xml',
    #              'light_blue.xml',
    #              'light_cyan.xml',
    #              'light_cyan_500.xml',
    #              'light_lightgreen.xml',
    #              'light_pink.xml',
    #              'light_purple.xml',
    #              'light_red.xml',
    #              'light_teal.xml',
    #              'light_yellow.xml']
    # apply_stylesheet(app, theme=themeList[10], invert_secondary=True, extra=extra)
    myWin = MyMainForm()
    helpWin = helpWindow()
    myWin.pushButton_10.clicked.connect(helpWin.show)
    myWin = CandyWindow.createWindow(myWin, 'blueDeep', title='副车架台架载荷迭代工具')

    myWin.show()

    sys.exit(app.exec_())
